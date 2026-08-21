#!/usr/bin/env python3
"""
LegalSmegal — Scotland £0 Valuation Band + Anchor Engine  (Slice 1, isolated prototype)

Doctrine: goal v1.1 / core-governance. Data -> Context -> Visuals -> User Inference.
NO fabricated values. Every output figure carries source + year + geography level + vintage.
Missing/suppressed data -> explicit unavailable state, never a guess.

ISOLATION: standalone module. Reads ONLY the four verified free Scottish files below.
It does NOT touch the E&W ceiling engine, comps tables, Supabase price-paid objects,
the live database, or any locked LegalSmegal subsystem. Integration into the deal
package (deal_id / Area Intelligence) is a separate, reviewed step — not done here.

VERIFIED INPUTS (all confirmed against the real files earlier):
  1. residential-properties-sales-and-price.csv  (statistics.gov.scot)
       Count/LQ/Mean/Median/UQ, annual 2004-2023, <5 suppressed, DATA ZONE 2011 vintage,
       multi-geography (S01 DZ, S02 IZ, S12 council, ...). This slice = Median 2023.
  2. SmallUser.csv  (NRS Scottish Postcode Directory, from spd_postcodeindex_cut)
       postcode -> DataZone2011Code, IntermediateZone2011Code, CouncilArea2019Code,
       SIMD2020 rank, UrbanRural6Fold, lat/long. Latin-1, 100% populated, live-filterable.
  3. ros_all_stats_June_2026.xlsx  (Registers of Scotland)
       M1 = monthly LA median (recency, to Jun 2026). C1 = annual LA median (2023 baseline).
       C2 = LA median by house type (type shape). LA is the finest RoS geography.
  4. Home Report PDF(s)  — per-lot anchor: Market Value + inspection/valuation date,
       property type, EPC, repair categories.
"""

import csv, re, subprocess, sys
from pathlib import Path

# ---- vintage boundaries (verified from the real files) -----------------------
DZ2011_MIN, DZ2011_MAX = "S01006506", "S01013481"   # 6,976 zones
N_DZ2011 = 6976                                       # for SIMD decile

# ---- required attribution (OGL v3.0) — MUST be displayed where data is shown --
ATTRIBUTION = [
    "Contains Registers of Scotland data © Crown copyright and database right 2026 (OGL v3.0).",
    "Contains National Records of Scotland data © Crown copyright (Scottish Postcode Directory, OGL v3.0).",
    "Contains Ordnance Survey data © Crown copyright and database right 2026.",
    "House price small-area statistics: Scottish Government / statistics.gov.scot (OGL v3.0).",
]

# ============================================================================
# REFERENCE DATA LOADERS  (fail loud; never fabricate a missing value)
# ============================================================================

def load_postcode_index(path):
    """postcode(no spaces, upper) -> geography dict. Live rows only."""
    idx = {}
    with open(path, newline="", encoding="latin-1") as f:
        for r in csv.DictReader(f):
            if r["DateOfDeletion"].strip():          # skip retired postcodes
                continue
            pc = r["Postcode"].replace(" ", "").upper()
            idx[pc] = {
                "dz2011": r["DataZone2011Code"].strip(),
                "iz2011": r["IntermediateZone2011Code"].strip(),
                "la_code": r["CouncilArea2019Code"].strip(),
                "simd2020_rank": _int(r["ScottishIndexOfMultipleDeprivation2020Rank"]),
                "ur6": r["UrbanRural6Fold2022Code"].strip(),
                "lat": _float(r["Latitude"]), "long": _float(r["Longitude"]),
            }
    return idx

def load_price_median_2023(path):
    """geography code -> median 2023 (int). Suppressed areas are simply absent."""
    out = {}
    with open(path, newline="", encoding="latin-1") as f:
        for row in csv.reader(f):
            if not row or "statistical-geography/" not in row[0]:
                continue
            code = row[0].rsplit("/", 1)[-1].strip()
            val = _int(row[-1]) if len(row) >= 3 else None
            if code.startswith(("S01", "S02", "S12", "S92")) and val:
                out[code] = val
    return out

def load_ros_la_medians(path):
    """Returns (current: la_code->median, current_period), (y2023: la_code->median)."""
    import openpyxl  # lazy: only the file/xlsx build path needs it, not the pg runtime
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    # M1 monthly LA -> latest median per LA
    cur, latest_period = {}, None
    ws = wb["M1"]                                          # leading blank col A -> data in cols 1-4
    rows = list(ws.iter_rows(min_row=7, values_only=True))
    for row in rows:
        month, la_name, la_code, median = row[1], row[2], row[3], row[4]
        if la_code and isinstance(median, (int, float)):
            cur[str(la_code).strip()] = int(median)      # last write = latest month
            latest_period = month
    # C1 annual LA -> 2023 median per LA
    y2023 = {}
    ws = wb["C1"]
    for row in ws.iter_rows(min_row=7, values_only=True):
        year, la_name, la_code, median = row[1], row[2], row[3], row[4]
        if str(year).strip() == "2023" and la_code and isinstance(median, (int, float)):
            y2023[str(la_code).strip()] = int(median)
    return cur, latest_period, y2023

def _int(x):
    try: return int(float(str(x).replace(",", "").strip()))
    except (ValueError, TypeError): return None

def _float(x):
    try: return float(str(x).strip())
    except (ValueError, TypeError): return None


# ============================================================================
# HOME REPORT ANCHOR EXTRACTOR  (structure-driven; returns None when absent)
# ============================================================================

PC_RE = re.compile(r"\b([A-Z]{1,2}\d{1,2}[A-Z]?)\s*(\d[A-Z]{2})\b")
# order matters: specific types before "detached" (which is a substring of "semi-detached")
TYPE_MAP = [("semi-detached", "semi"), ("semi detached", "semi"),
            ("end terraced", "terraced"), ("mid terraced", "terraced"), ("terraced", "terraced"),
            ("flat", "flat"), ("maisonette", "flat"), ("bungalow", "bungalow"),
            ("detached", "detached"), ("villa", "house")]

def extract_home_report_from_text(txt, source_file=""):
    """Regex the anchor fields out of already-extracted Home Report text.
    Same field logic as extract_home_report; any field not found -> None.
    This is the single source of truth for the anchor regexes — the PDF entry
    point calls it after pdftotext, and the app calls it on stored document text."""
    a = {"source_file": source_file, "market_value": None, "valuation_date": None,
         "property_type": None, "epc_band": None, "cat3_count": None, "postcode": None,
         "error": None}
    if not txt or not txt.strip():
        a["error"] = "no extractable text (scanned or encrypted pack?)"
        return a

    m = PC_RE.search(txt)                                     # postcode
    if m: a["postcode"] = (m.group(1) + m.group(2)).upper()

    # anchor value: Format A (MVR table) OR Format B (Single Survey narrative).
    # NB "not less than £X" / "insured ... £X" is the REINSTATEMENT figure -> never market value.
    m = re.search(r"Market value in present condition\s*:?\s*£?\s*([\d,]{5,})", txt, re.I)
    if not m:  # "fairly stated in the region of/sum of" = market value; "not less than" = reinstatement (excluded)
        m = re.search(r"fairly stated in (?:the region of|the sum of)\s*£?\s*([\d,]{5,})", txt, re.I)
    if m: a["market_value"] = _int(m.group(1))

    m = re.search(r"Date of Inspection\s+([0-9]{1,2}[a-z]{0,2}\s+\w+\s+\d{4})", txt, re.I)
    if not m:
        m = re.search(r"as at (?:the\s+)?([0-9]{1,2}[a-z]{0,2}\s+\w+\s+\d{4})", txt, re.I)
    if m: a["valuation_date"] = m.group(1).strip()

    # property type: classify from the "subjects comprise ..." description ONLY
    low = txt.lower()
    desc = ""
    md = re.search(r"(?:subjects? comprises?|comprises? an?)(.{0,90})", low)
    if md: desc = md.group(1)
    for needle, norm in TYPE_MAP:
        if needle in desc:
            a["property_type"] = norm; break

    m = re.search(r"band\s+([A-G])\b", txt)                   # EPC band (first occurrence)
    if m: a["epc_band"] = m.group(1)

    a["cat3_count"] = len(re.findall(r"Repair category\s+3\b", txt))  # urgent-repair flags
    return a


def extract_home_report(pdf_path, max_bytes=25_000_000, timeout=20):
    """Pull anchor fields from a Home Report PDF. Any field not found -> None.
    Hardened for untrusted uploads: size guard, subprocess timeout, graceful failure.
    On failure returns the dict with error set and fields None (never raises, never guesses)."""
    src_name = Path(pdf_path).name
    _err = lambda msg: {"source_file": src_name, "market_value": None, "valuation_date": None,
                        "property_type": None, "epc_band": None, "cat3_count": None,
                        "postcode": None, "error": msg}
    try:
        if Path(pdf_path).stat().st_size > max_bytes:
            return _err("file too large")
        proc = subprocess.run(["pdftotext", "-layout", str(pdf_path), "-"],
                              capture_output=True, text=True, timeout=timeout)
        txt = proc.stdout
    except (subprocess.TimeoutExpired, FileNotFoundError, OSError) as ex:
        return _err(f"pdf read failed: {type(ex).__name__}")
    return extract_home_report_from_text(txt, src_name)


# ============================================================================
# BAND BUILDER  (tiered: DZ2011 -> IZ2011 -> LA floor; each labelled + sourced)
# ============================================================================

def simd_decile(rank):
    if not rank: return None
    return ((rank - 1) * 10 // N_DZ2011) + 1        # 1 = most deprived

def build_band(postcode, ref):
    """Return a fully-sourced band dict, or an explicit unavailable state."""
    pc = (postcode or "").replace(" ", "").upper()
    geo = ref["pc_index"].get(pc)
    if not geo:
        return {"status": "unavailable", "reason": f"postcode {postcode} not in live SPD index"}

    price = ref["price_2023"]
    # tier selection: finest available wins
    tier = None
    if geo["dz2011"] in price:
        tier = ("Data Zone (2011)", geo["dz2011"], price[geo["dz2011"]])
    elif geo["iz2011"] in price:
        tier = ("Intermediate Zone (2011)", geo["iz2011"], price[geo["iz2011"]],
                "data zone suppressed (<5 sales)")
    elif geo["la_code"] in price:
        tier = ("Local Authority", geo["la_code"], price[geo["la_code"]],
                "data zone & intermediate zone suppressed")
    if tier is None:
        return {"status": "unavailable",
                "reason": "no unsuppressed area figure at DZ/IZ/LA for median 2023"}

    band_2023 = tier[2]
    # time-adjust 2023 -> current using RoS LA-level growth (same-source factor)
    la = geo["la_code"]
    factor, adj = None, None
    if la in ref["ros_cur"] and la in ref["ros_2023"] and ref["ros_2023"][la]:
        factor = ref["ros_cur"][la] / ref["ros_2023"][la]
        adj = round(band_2023 * factor)

    return {
        "status": "ok",
        "geography_level": tier[0],
        "area_code": tier[1],
        "band_median_2023": band_2023,
        "band_source": "statistics.gov.scot residential-properties-sales-and-price (RoS data)",
        "tier_note": tier[3] if len(tier) > 3 else None,
        "time_adjust_factor": round(factor, 4) if factor else None,
        "band_median_current": adj,
        "time_adjust_source": ("RoS LA median {}/2023 -> {}".format(la, ref["ros_period"])
                               if factor else "unavailable (LA not matched in RoS)"),
        "la_current_median": ref["ros_cur"].get(la),        # recency cross-check tier
        "la_current_period": ref["ros_period"],
        "simd2020_decile": simd_decile(geo["simd2020_rank"]),
        "urban_rural_6fold": geo["ur6"],
        "lat": geo["lat"], "long": geo["long"],
    }


# ============================================================================
# ASSEMBLE  (anchor + band + honesty states + divergence flag)
# ============================================================================

def assess_lot(pdf_path, ref, divergence_pct=40):
    """
    divergence_pct is a PLACEHOLDER threshold for the prototype. Per the spec it is
    [CALIBRATE] against real lots in MEASURE->L4; it is NOT a claimed production value.
    """
    anchor = extract_home_report(pdf_path)
    band = build_band(anchor["postcode"], ref)
    out = {"anchor": anchor, "band": band, "flags": []}

    if band["status"] != "ok":
        out["flags"].append("BAND UNAVAILABLE: " + band["reason"])
        return out
    if anchor["market_value"] is None:
        out["flags"].append("NO HOME REPORT VALUATION FOUND — band shown alone")
        return out

    # anchor vs band divergence (compare like-for-like: current-adjusted band if available)
    ref_band = band["band_median_current"] or band["band_median_2023"]
    if ref_band:
        div = (anchor["market_value"] - ref_band) / ref_band * 100
        out["divergence_pct"] = round(div, 1)
        if abs(div) >= divergence_pct:
            out["flags"].append(
                f"ANCHOR vs BAND DIVERGENCE {div:+.0f}% "
                f"(HR £{anchor['market_value']:,} vs area £{ref_band:,}) — surface both")
    if anchor["cat3_count"]:
        out["flags"].append(f"{anchor['cat3_count']} element(s) at Repair Category 3 (urgent)")
    return out


# ============================================================================
# DEMO RUN  (real data, five real Home Reports)
# ============================================================================

def load_reference(ref_dir):
    ref_dir = Path(ref_dir)
    cur, period, y2023 = load_ros_la_medians(ref_dir / "ros_all_stats_June_2026.xlsx")
    return {
        "pc_index":   load_postcode_index(ref_dir / "SmallUser.csv"),
        "price_2023": load_price_median_2023(ref_dir / "residential-properties-sales-and-price.csv"),
        "ros_cur": cur, "ros_period": period, "ros_2023": y2023,
    }

def _fmt(v): return f"£{v:,}" if isinstance(v, int) else str(v)

if __name__ == "__main__":
    ref_dir = sys.argv[1] if len(sys.argv) > 1 else "."
    hr_dir  = sys.argv[2] if len(sys.argv) > 2 else "."
    ref = load_reference(ref_dir)
    print(f"reference loaded: {len(ref['pc_index']):,} live postcodes | "
          f"{len(ref['price_2023']):,} area price rows | "
          f"RoS current period: {ref['ros_period']}\n")

    for pdf in sorted(Path(hr_dir).glob("*.pdf")):
        r = assess_lot(pdf, ref)
        a, b = r["anchor"], r["band"]
        print("=" * 78)
        print(f"{a['postcode']}  | type: {a['property_type']} | EPC {a['epc_band']} "
              f"| HR value {_fmt(a['market_value'])} (as at {a['valuation_date']})")
        if b["status"] == "ok":
            print(f"  band: {b['geography_level']} {b['area_code']} "
                  f"median 2023 {_fmt(b['band_median_2023'])}"
                  + (f" -> current {_fmt(b['band_median_current'])} (x{b['time_adjust_factor']})"
                     if b['band_median_current'] else " -> current: unavailable"))
            print(f"  recency: LA current median {_fmt(b['la_current_median'])} ({b['la_current_period']}) "
                  f"| SIMD2020 decile {b['simd2020_decile']} | urban-rural {b['urban_rural_6fold']}")
            if b["tier_note"]: print(f"  tier note: {b['tier_note']}")
            if "divergence_pct" in r: print(f"  divergence: {r['divergence_pct']:+}%")
        for fl in r["flags"]:
            print(f"  ⚑ {fl}")
    print("=" * 78)
